import openpyxl


def get_first_two_columns(filename):
    workbook = openpyxl.load_workbook(f'downloads/{filename}')

    sheet = workbook.active

    data = []

    for row in sheet.iter_rows():
        first_column = row[0].value
        second_column = row[1].value

        data.append([first_column, second_column])

    return data


def create_book(name, lis: list):
    filepath = name
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet["A1"].value = "фио"
    sheet["B1"].value = "Телефоны"
    for dic in lis:
        model = dic["model"]
        number = dic["number"]
        year = dic["year"]

        salt = dic['salt']
        if len(dic['phone']) != 0:

            for i in range(len(dic['phone'])):
                sheet.append((f"{number} {model} {year} {salt} {i}",dic['phone'][i]))

    wb.save(filepath, )
text=str("🇷🇺 Страна: Россия\
\
🔎 Всего результатов: 2\
🧍‍♂️ Личности: АГАПОВА ЕЛЕНА ГЕННАДЬЕВНА 09.07.1974\
🚘 Транспорт: О864УН96")
first = text.find("Транспорт:") + 11
last = text.find("Транспорт: ") + 19
number = text[first:last].upper().replace("\n", "")
print(number)

